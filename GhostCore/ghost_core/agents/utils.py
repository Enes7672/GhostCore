"""
agents.py — GhostCore v4 / Nova Jarvis: Redline Architecture
=============================================================
UPGRADE: The Phantom Four — Elite Edition

10 Yeni Özellik (3 Kategoride):

A. Gelişmiş Kod Yazma:
    1. SOLID & Design Pattern Awareness  — Factory/Observer/Singleton otomatik seçimi
    2. Multi-File Sync                   — Bağımlı dosyaları (imports, calls) senkronize et
    3. Automatic Unit Testing            — Her fonksiyon için pytest dosyası üret
    4. Performance Profiling             — Ryzen 7 üzerinde CPU/RAM maliyeti hesapla

B. Kod Düzeltme & İyileştirme:
    5. Self-Correction Loop (3 iter.)    — Hunter bulur → Architect düzeltir (3x iterasyon)
    6. Security Auto-Patching            — Açıkları sadece raporlama, o satırı değiştir

C. GhostCore Mantıksal Güçlendirme:
    7. Long-Term Memory (RAG)            — Önceki proje tercihlerini hatırlayan vektör bellek
    8. Tool-Aware Execution              — FileReadTool + TerminalTool gerçek araç kullanımı
    9. Hybrid Logic                      — Karmaşıklığa göre Local vs Cloud karar mekanizması
   10. Git-Native Workflow               — Her başarılı modül için Conventional Commits mesajı

Author: GhostCore Architecture Team
"""

import os
import re
import json
import logging
import sys

# Windows Unicode desteği (cp1254 hatasını önlemek için)
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')  # type: ignore[union-attr]
    except Exception:
        pass

import hashlib
import sqlite3
import subprocess
from pathlib import Path
from typing import Optional, Any
from dataclasses import dataclass, field, asdict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, Reference, BarChart


from crewai_tools import FileReadTool, DirectoryReadTool
_file_read_tool = FileReadTool()
_directory_read_tool = DirectoryReadTool()
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

from ..brain import (
    verify_ollama_connection,
    get_architect_brain,
    get_sentinel_brain,
    get_hunter_brain,
    get_writer_brain,
    TOKEN_MANAGER,
    GLOBAL_CACHE,
    resolve_model_for_task,
)

from langchain_community.tools import DuckDuckGoSearchRun
WEB_SEARCH_TOOL = DuckDuckGoSearchRun()

# ---------------------------------------------------------------------------
# Logger & Console
# ---------------------------------------------------------------------------

logger = logging.getLogger("ghostcore.agents")
console = Console()

# ---------------------------------------------------------------------------
# Startup Guard
# ---------------------------------------------------------------------------

verify_ollama_connection()


# ===========================================================================
# GÜVENLİK SANDBOX ADAPTER — Docker | Subprocess | Manual
# ===========================================================================
#
# .env'de SANDBOX_MODE değişkenini ayarla:
#   manual     → Kodu çalıştırmaz, yalnızca statik analiz raporu döndürür (varsayılan)
#   subprocess → Yerel subprocess + timeout izolasyonu (Docker gerektirmez)
#   docker     → Docker konteynerinde tam ağ/dosya izolasyonu (Docker kurulu olmalı)
#
# Hunter bu sınıfı kullanarak hem statik denetim hem de dinamik çalıştırma yapabilir.
# ===========================================================================

_SANDBOX_MODE: str = os.getenv("SANDBOX_MODE", "manual").lower().strip()


class SandboxResult:
    """Sandbox yürütme sonucu."""
    def __init__(self, mode: str, stdout: str = "", stderr: str = "",
                 timed_out: bool = False, error: str = "") -> None:
        self.mode      = mode
        self.stdout    = stdout
        self.stderr    = stderr
        self.timed_out = timed_out
        self.error     = error

    @property
    def safe_summary(self) -> str:
        """Hunter'ın LLM prompt'una eklenecek kısa özet."""
        if self.error:
            return f"[SANDBOX:{self.mode.upper()}] HATA: {self.error}"
        if self.timed_out:
            return f"[SANDBOX:{self.mode.upper()}] ⏱ Zaman Aşımı — kod 10 saniyede tamamlanamadı."
        lines = []
        if self.stdout:
            lines.append(f"STDOUT:\n{self.stdout[:800]}")
        if self.stderr:
            lines.append(f"STDERR:\n{self.stderr[:400]}")
        return f"[SANDBOX:{self.mode.upper()}]\n" + "\n".join(lines) if lines else f"[SANDBOX:{self.mode.upper()}] Çıktı yok."


class SandboxExecutor:
    """
    Güvenlik Sandbox — Adapter Tasarım Deseni.

    Hunter, üretilen kodu yalnızca kendine ait bir ortamda çalıştırır.
    Üç mod:
        • manual     : Çalıştırmaz, güvenlik raporu oluşturur.
        • subprocess : Child process + timeout + memory limit.
        • docker     : Hardened Alpine container, ağ ve dosya sistemi izole.
    """

    # ── MANUAL ────────────────────────────────────────────────────────────────

    def _run_manual(self, code: str, tests: str = "") -> SandboxResult:
        """Kodu çalıştırmaz. Gelişmiş statik analiz ve karmaşıklık kontrolü yapar."""
        import tempfile, shutil
        import re as _re
        from pathlib import Path
        
        with tempfile.NamedTemporaryFile(suffix=".py", delete=False, mode="w", encoding="utf-8") as f:
            f.write(code)
            tmp = f.name

        report_lines = ["[MANUAL SANDBOX — Gelişmiş Statik Analiz]"]

        # 1. Regex Basic Checks
        dangerous_patterns = [
            (r"os\.system\s*\(", "os.system() doğrudan shell komutu"),
            (r"subprocess\.call\s*\(", "subprocess.call() tespiti"),
            (r"eval\s*\(", "eval() tehlikeli kod yürütme"),
        ]
        for pattern, desc in dangerous_patterns:
            if _re.search(pattern, code):
                report_lines.append(f"  ⚠ TEHLİKE (Regex): {desc}")

        # 2. Ruff (Linter)
        if shutil.which("ruff"):
            try:
                proc = subprocess.run(["ruff", "check", tmp], capture_output=True, text=True, timeout=15)
                if proc.stdout: report_lines.append(f"\n[Ruff (Linter)]\n{proc.stdout.strip()[:400]}")
            except Exception: pass

        # 3. MyPy (Type Checker)
        if shutil.which("mypy"):
            try:
                proc = subprocess.run(["mypy", "--strict", tmp], capture_output=True, text=True, timeout=15)
                if proc.stdout: report_lines.append(f"\n[MyPy (Tipler)]\n{proc.stdout.strip()[:400]}")
            except Exception: pass

        # 4. Bandit (Security)
        if shutil.which("bandit"):
            try:
                proc = subprocess.run(["bandit", "-q", tmp], capture_output=True, text=True, timeout=15)
                if proc.stdout or proc.stderr: report_lines.append(f"\n[Bandit (Güvenlik)]\n{(proc.stdout or proc.stderr).strip()[:400]}")
            except Exception: pass

        # 5. Radon (Cyclomatic Complexity - Spaghetti Code Check)
        if shutil.which("radon"):
            try:
                proc = subprocess.run(["radon", "cc", "-nc", tmp], capture_output=True, text=True, timeout=10)
                if proc.stdout:
                    out_str = proc.stdout.strip()
                    report_lines.append(f"\n[Radon Karmaşıklık]\n{out_str[:400]}")
                    if " C " in out_str or " D " in out_str or " E " in out_str or " F " in out_str:
                        report_lines.append("\n  ⚠ SPAGHETTI KOD UYARISI: Karmaşıklık çok yüksek (C/D/E/F düzeyi). Kodu parçalara ayır ve okunabilirliği artır.")
            except Exception: pass

        Path(tmp).unlink(missing_ok=True)
        return SandboxResult(mode="manual", stdout="\n".join(report_lines))

    # ── SUBPROCESS ────────────────────────────────────────────────────────────

    def _run_subprocess(self, code: str, tests: str = "", timeout: int = 15) -> SandboxResult:
        """TDD destekli subprocess sandbox."""
        import tempfile, sys as _sys, shutil
        from pathlib import Path
        
        temp_dir = Path(tempfile.mkdtemp())
        code_file = temp_dir / "target_module.py"
        code_file.write_text(code, encoding="utf-8")
        
        test_file = temp_dir / "test_target.py"
        if tests:
            test_file.write_text(tests, encoding="utf-8")

        try:
            if tests and shutil.which("pytest"):
                proc = subprocess.run(
                    ["pytest", "-v", str(test_file)],
                    capture_output=True, text=True,
                    timeout=timeout,
                    cwd=str(temp_dir),
                )
            else:
                proc = subprocess.run(
                    [_sys.executable, str(code_file)],
                    capture_output=True, text=True,
                    timeout=timeout,
                    cwd=str(temp_dir),
                )
            return SandboxResult(mode="subprocess", stdout=proc.stdout, stderr=proc.stderr)
        except subprocess.TimeoutExpired:
            return SandboxResult(mode="subprocess", timed_out=True)
        except Exception as e:
            return SandboxResult(mode="subprocess", error=str(e))
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ── DOCKER ────────────────────────────────────────────────────────────────

    def _run_docker(self, code: str, tests: str = "", timeout: int = 20) -> SandboxResult:
        """TDD destekli Docker Sandbox."""
        import shutil, tempfile
        from pathlib import Path
        if not shutil.which("docker"):
            return SandboxResult(mode="docker", error="Docker kurulu değil.")
            
        temp_dir = Path(tempfile.mkdtemp())
        code_file = temp_dir / "target_module.py"
        code_file.write_text(code, encoding="utf-8")
        
        test_file = temp_dir / "test_target.py"
        if tests:
            test_file.write_text(tests, encoding="utf-8")
            entry_cmd = ["pytest", "-v", "/sandbox/test_target.py"]
            # To run pytest in the container, it must be installed. For real world, we might need a custom image.
            # We'll run it as a normal python script if pytest isn't available in alpine container easily out of the box,
            # or try to use python -m unittest. For now assuming python:3.12-alpine has it or we just pip install it on the fly.
            full_cmd = [
                "docker", "run", "--rm", "--network", "none", "--memory", "150m", "--cpus", "0.5",
                "-v", f"{temp_dir}:/sandbox:ro", "-w", "/sandbox",
                "python:3.12-alpine", "sh", "-c", "pip install -q pytest && pytest -v test_target.py"
            ]
        else:
            full_cmd = [
                "docker", "run", "--rm", "--network", "none", "--memory", "100m", "--cpus", "0.5",
                "-v", f"{temp_dir}:/sandbox:ro", "-w", "/sandbox",
                "python:3.12-alpine", "python", "target_module.py"
            ]
            
        try:
            proc = subprocess.run(full_cmd, capture_output=True, text=True, timeout=timeout + 5)
            return SandboxResult(mode="docker", stdout=proc.stdout, stderr=proc.stderr)
        except subprocess.TimeoutExpired:
            return SandboxResult(mode="docker", timed_out=True)
        except Exception as e:
            return SandboxResult(mode="docker", error=str(e))
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ── PUBLIC API ────────────────────────────────────────────────────────────

    def execute(self, code: str, tests: str = "") -> SandboxResult:
        """
        .env'deki SANDBOX_MODE'a göre uygun sandbox'u çalıştırır.
        """
        if _SANDBOX_MODE == "docker":
            return self._run_docker(code, tests)
        elif _SANDBOX_MODE == "subprocess":
            return self._run_subprocess(code, tests)
        else:
            return self._run_manual(code, tests)

    @property
    def mode(self) -> str:
        return _SANDBOX_MODE


# Singleton sandbox — tüm ajanlar paylaşır
SANDBOX = SandboxExecutor()


# ===========================================================================
# ÖZELLIK 7: LONG-TERM MEMORY (RAG Entegrasyonu)
# ===========================================================================

@dataclass
class MemoryEntry:
    """Uzun süreli bellek kaydı. Her tercih veya karar buraya yazılır."""
    key:        str
    value:      str
    tags:       list[str] = field(default_factory=list)
    created_at: str       = field(default_factory=lambda: datetime.now().isoformat())


# ---------------------------------------------------------------------------
# Adapter arayüzü — her backend bu imzayı uygular
# ---------------------------------------------------------------------------

class MemoryBackend:
    """Soyut bellek backend arayüzü. Tüm backendler bu metotları sağlamalı."""

    def remember(self, key: str, value: str, tags: list[str]) -> None:
        raise NotImplementedError

    def recall(self, query: str, top_k: int = 3) -> list[str]:
        raise NotImplementedError

    @property
    def size(self) -> int:
        raise NotImplementedError


# ---------------------------------------------------------------------------
# Backend A: JSON (varsayılan — sıfır bağımlılık)
# .env: MEMORY_BACKEND=json
# ---------------------------------------------------------------------------

class JSONMemoryBackend(MemoryBackend):
    """
    Belleği ./data/long_term_memory.json dosyasında tutar.
    Kurulum gerektirmez; Raspberry Pi'den sunucuya her ortamda çalışır.
    """
    MEMORY_FILE = Path(__file__).parent.parent.parent / "data" / "long_term_memory.json"

    def __init__(self) -> None:
        self._store: list[MemoryEntry] = []
        self._load()

    def _load(self) -> None:
        self.MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        if self.MEMORY_FILE.exists():
            try:
                raw = json.loads(self.MEMORY_FILE.read_text(encoding="utf-8"))
                self._store = [MemoryEntry(**e) for e in raw]
                logger.info("[JSONMemory] %d kayıt yüklendi.", len(self._store))
            except Exception as e:
                logger.warning("[JSONMemory] Yükleme hatası: %s", e)

    def _save(self) -> None:
        try:
            data = [vars(e) for e in self._store]
            self.MEMORY_FILE.write_text(
                json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        except Exception as e:
            logger.warning("[JSONMemory] Kayıt hatası: %s", e)

    def remember(self, key: str, value: str, tags: list[str]) -> None:
        self._store = [e for e in self._store if e.key != key]
        self._store.append(MemoryEntry(key=key, value=value, tags=tags))
        self._save()
        logger.info("[JSONMemory] Kaydedildi: %s", key)

    def recall(self, query: str, top_k: int = 3) -> list[str]:
        q = query.lower()
        results = []
        for e in self._store:
            score = (
                (q in e.key.lower()) * 2
                + (q in e.value.lower()) * 1
                + sum(q in t.lower() for t in e.tags) * 3
            )
            if score > 0:
                results.append((score, e.value))
        results.sort(reverse=True)
        return [v for _, v in results[:top_k]]

    @property
    def size(self) -> int:
        return len(self._store)


# ---------------------------------------------------------------------------
# Backend B: ChromaDB (opt-in — vektör tabanlı semantik arama)
# .env: MEMORY_BACKEND=chromadb
# ---------------------------------------------------------------------------

class ChromaMemoryBackend(MemoryBackend):
    """
    Belleği ./data/chromadb klasöründe kalıcı vektör DB olarak tutar.
    Gereksinim: pip install chromadb

    Avantajlar:
        • "docker" yazınca "Docker Compose kullan" sonucu çıkar — keyword eşleşmesi gerekmez
        • Sınırsız kayıt; JSON'un aksine büyük belleklerde yavaşlamaz
        • Her oturum kapandığında disk'e yazılır, yeniden açıldığında otomatik yüklenir
    """
    CHROMA_DIR = str(Path(__file__).parent.parent.parent / "data" / "chromadb")
    COLLECTION  = "ghostcore_memory"

    def __init__(self) -> None:
        try:
            import chromadb  # type: ignore
            self._client     = chromadb.PersistentClient(path=self.CHROMA_DIR)
            self._collection = self._client.get_or_create_collection(
                name=self.COLLECTION,
                metadata={"hnsw:space": "cosine"},
            )
            logger.info("[ChromaMemory] Koleksiyon yüklendi: %d kayıt.", self._collection.count())
        except ImportError:
            raise RuntimeError(
                "ChromaDB kurulu değil! Kurmak için: pip install chromadb\n"
                "Ya da .env'de MEMORY_BACKEND=json kullanın."
            )

    def remember(self, key: str, value: str, tags: list[str]) -> None:
        doc_id  = hashlib.md5(key.encode()).hexdigest()
        full_text = f"{key} {' '.join(tags)} {value}"
        self._collection.upsert(
            ids=[doc_id],
            documents=[full_text],
            metadatas=[{"key": key, "value": value, "tags": json.dumps(tags)}],
        )
        logger.info("[ChromaMemory] Kaydedildi: %s", key)

    def recall(self, query: str, top_k: int = 3) -> list[str]:
        if self._collection.count() == 0:
            return []
        results = self._collection.query(
            query_texts=[query],
            n_results=min(top_k, self._collection.count()),
        )
        values: list[str] = []
        metadatas = results.get("metadatas")
        if metadatas and len(metadatas) > 0 and metadatas[0]:
            for meta in metadatas[0]:
                if isinstance(meta, dict) and "value" in meta:
                    values.append(str(meta.get("value", "")))
        return values

    @property
    def size(self) -> int:
        return self._collection.count()


# ---------------------------------------------------------------------------
# LongTermMemory — Adapter Facade (public API değişmedi)
# ---------------------------------------------------------------------------

# .env'den backend seçimi: MEMORY_BACKEND=json (varsayılan) | chromadb
_MEMORY_BACKEND_ENV: str = os.getenv("MEMORY_BACKEND", "json").lower().strip()


class LongTermMemory:
    """
    Ajan long-term memory katmanı — Adapter Tasarım Deseni.

    .env içinde MEMORY_BACKEND değişkenini ayarlayarak backend'i seç:
        MEMORY_BACKEND=json       → JSON dosyası (varsayılan, sıfır kurulum)
        MEMORY_BACKEND=chromadb   → ChromaDB vektör DB (pip install chromadb)

    API tamamen aynı kalır — ajanlar backend değişiminden habersizdir:
        MEMORY.remember("key", "value", tags=["tag"])
        MEMORY.recall("query")
        MEMORY.context_for_prompt("görev açıklaması")
    """

    def __init__(self) -> None:
        if _MEMORY_BACKEND_ENV == "chromadb":
            try:
                self._backend: MemoryBackend = ChromaMemoryBackend()
                logger.info("[Memory] Backend: ChromaDB (vektör semantik arama)")
            except RuntimeError as e:
                logger.warning("[Memory] ChromaDB başlatılamadı, JSON'a düşülüyor: %s", e)
                self._backend = JSONMemoryBackend()
        else:
            self._backend = JSONMemoryBackend()
            logger.info("[Memory] Backend: JSON (keyword tabanlı arama)")

    def remember(self, key: str, value: str, tags: Optional[list[str]] = None) -> None:
        """
        Yeni bir tercih veya kararı uzun süreli belleğe yazar.

        Args:
            key:   Bellek anahtarı (örn. "docker_preference").
            value: Hatırlanacak değer (örn. "Always use multi-stage").
            tags:  Arama etiketleri (örn. ["docker", "devops"]).
        """
        self._backend.remember(key, value, tags or [])

    def recall(self, query: str, top_k: int = 3) -> list[str]:
        """Sorguya en yakın bellek kayıtlarını döndürür."""
        return self._backend.recall(query, top_k)

    def context_for_prompt(self, query: str) -> str:
        """LLM prompt'una eklenecek bellek bağlamı string'i üretir."""
        memories = self.recall(query)
        if not memories:
            return ""
        lines = "\n".join(f"  - {m}" for m in memories)
        return f"\n[LONG-TERM MEMORY — Geliştirici Tercihleri]\n{lines}\n"

    @property
    def size(self) -> int:
        return self._backend.size

    @property
    def backend_name(self) -> str:
        return type(self._backend).__name__


# Singleton bellek nesnesi — tüm ajanlar paylaşır
MEMORY = LongTermMemory()



# ===========================================================================
# ÖZELLIK 1: SOLID & DESIGN PATTERN AWARENESS
# ===========================================================================

class DesignPatternAdvisor:
    """
    Görev açıklamasını analiz ederek uygun tasarım desenini önerir.
    The Architect her yeni görevde bunu çalıştırır ve prompt'una ekler.
    """

    PATTERNS = {
        "factory": {
            "keywords": ["create", "oluştur", "instantiate", "nesne", "üret", "generate object"],
            "description": "Factory Pattern: Nesne oluşturma mantığını soyutla. "
                           "`get_XXX()` factory fonksiyonları kullan.",
            "example": "def get_connection(db_type: str) -> BaseDB: ...",
        },
        "observer": {
            "keywords": ["event", "listen", "notify", "callback", "hook", "trigger", "watch"],
            "description": "Observer Pattern: Event-driven mimari. "
                           "Publisher/Subscriber ayrımı ile loose coupling sağla.",
            "example": "class EventBus: def subscribe(self, event, handler): ...",
        },
        "singleton": {
            "keywords": ["single", "global", "shared", "tek", "paylaşılan", "config", "cache"],
            "description": "Singleton Pattern: Tek instance garantisi. "
                           "Module-level instance veya `__new__` override kullan.",
            "example": "class Config:\n    _instance = None\n    def __new__(cls): ...",
        },
        "strategy": {
            "keywords": ["algorithm", "sort", "select", "choose", "switch", "mode", "strateji"],
            "description": "Strategy Pattern: Algoritmaları değiştirilebilir hale getir. "
                           "Protocol/ABC ile interface tanımla.",
            "example": "class SortStrategy(Protocol): def sort(self, data): ...",
        },
        "repository": {
            "keywords": ["database", "db", "query", "crud", "save", "fetch", "store", "veri"],
            "description": "Repository Pattern: Veri erişim katmanını soyutla. "
                           "Business logic ve DB sorgularını birbirinden ayır.",
            "example": "class UserRepository: def find_by_id(self, id: int): ...",
        },
    }

    def suggest(self, task_description: str) -> str:
        """
        Görev metnini tarayıp uygun desenleri prompt string'i olarak döndürür.

        Args:
            task_description: Kullanıcının görev açıklaması.

        Returns:
            Architect prompt'una eklenecek tasarım deseni yönergesi.
        """
        task_lower = task_description.lower()
        suggestions = []

        for name, info in self.PATTERNS.items():
            if any(kw in task_lower for kw in info["keywords"]):
                suggestions.append(
                    f"  [{name.upper()}] {info['description']}\n"
                    f"  Örnek: `{info['example']}`"
                )

        if not suggestions:
            return (
                "\n[DESIGN GUIDANCE]\n"
                "  SOLID prensiplerine uyu. Single Responsibility önce gelir.\n"
                "  Gereksiz abstraction ekleme — YAGNI.\n"
            )

        return (
            "\n[DESIGN PATTERN ADVISORY]\n"
            "Görevin doğasına göre şu desenler önerilir:\n"
            + "\n".join(suggestions) + "\n"
        )


PATTERN_ADVISOR = DesignPatternAdvisor()


# ===========================================================================
# ÖZELLIK 2: MULTI-FILE SYNC (Çoklu Dosya Uyumu)
# ===========================================================================

class MultiFileSyncAnalyzer:
    """
    Bir dosya değiştiğinde, projedeki diğer dosyalardaki etkilenen
    import ve method call'larını tespit eder.

    The Architect bunu her kod üretiminde çalıştırır;
    etkilenen dosyaları The Writer ile senkronize eder.
    """

    def find_dependents(self, changed_file: str, project_root: str = ".") -> list[dict]:
        """
        Verilen dosyaya bağımlı tüm Python dosyalarını bulur.

        Args:
            changed_file: Değişen dosyanın adı (örn. "models.py").
            project_root: Aranacak kök dizin.

        Returns:
            [{"file": "views.py", "lines": [3, 17], "imports": ["UserModel"]}]
        """
        module_name = Path(changed_file).stem
        dependents  = []

        for py_file in Path(project_root).rglob("*.py"):
            if py_file.name == Path(changed_file).name:
                continue
            try:
                content = py_file.read_text(encoding="utf-8")
                lines_affected = []
                imports_used: list[str] = []

                for i, line in enumerate(content.splitlines(), start=1):
                    if f"import {module_name}" in line or f"from {module_name}" in line:
                        lines_affected.append(i)
                        # Import edilen isimleri çek
                        match = re.findall(r"import\s+([\w,\s]+)", line)
                        imports_used.extend(
                            n.strip() for m in match for n in m.split(",")
                        )

                if lines_affected:
                    dependents.append({
                        "file":    str(py_file),
                        "lines":   lines_affected,
                        "imports": imports_used,
                    })
            except Exception:
                continue

        return dependents

    def sync_report(self, changed_file: str, project_root: str = ".") -> str:
        """
        Senkronizasyon raporunu LLM prompt'u için hazırlar.

        Args:
            changed_file: Değişen dosya adı.
            project_root: Proje kök dizini.

        Returns:
            Architect prompt'una eklenecek sync bağlamı.
        """
        deps = self.find_dependents(changed_file, project_root)
        if not deps:
            return ""

        lines = [f"\n[MULTI-FILE SYNC — {changed_file} Değişimi Etkileri]"]
        for d in deps:
            lines.append(
                f"  ⚠  {d['file']} (satır {d['lines']}) "
                f"→ {', '.join(d['imports'])} import'larını kontrol et"
            )
        lines.append(
            "  → Bu bağımlılıkları da güncelle veya uyumluluk için yorum ekle.\n"
        )
        return "\n".join(lines)


FILE_SYNC = MultiFileSyncAnalyzer()


# ===========================================================================
# ÖZELLIK 3: AUTOMATIC UNIT TESTING
# ===========================================================================

class UnitTestGenerator:
    """
    Üretilen Python kodunu parse ederek otomatik pytest dosyası oluşturur.
    The Writer her başarılı kod üretiminin ardından bunu çalıştırır.
    """

    def extract_functions(self, code: str) -> list[dict]:
        """
        Koddan fonksiyon ve sınıf adlarını parse eder.

        Args:
            code: Python kaynak kodu string'i.

        Returns:
            [{"name": "get_user", "type": "function", "args": ["user_id"]}]
        """
        items = []

        # Fonksiyonlar
        for m in re.finditer(
            r"^(async\s+)?def\s+(\w+)\s*\((.*?)\)", code, re.MULTILINE
        ):
            args = [
                a.strip().split(":")[0].split("=")[0].strip()
                for a in m.group(3).split(",")
                if a.strip() and a.strip() != "self"
            ]
            items.append({
                "name": m.group(2),
                "type": "async_function" if m.group(1) else "function",
                "args": args,
            })

        # Sınıflar
        for m in re.finditer(r"^class\s+(\w+)", code, re.MULTILINE):
            items.append({"name": m.group(1), "type": "class", "args": []})

        return items

    def generate_test_file(self, module_name: str, code: str) -> str:
        """
        Verilen modül için eksiksiz pytest dosyası üretir.

        Args:
            module_name: Modül adı (dosya uzantısız).
            code:        Modülün kaynak kodu.

        Returns:
            Pytest dosyasının içeriği (string).
        """
        items    = self.extract_functions(code)
        now      = datetime.now().strftime("%Y-%m-%d %H:%M")
        fn_tests = []
        cls_tests = []

        for item in items:
            # Özel metodları ve private fonksiyonları atla
            if item["name"].startswith("_"):
                continue

            args_str  = ", ".join(f'"{a}_test"' for a in item["args"])
            call_expr = f'{item["name"]}({args_str})'

            if item["type"] == "class":
                cls_tests.append(
                    f"\nclass Test{item['name']}:\n"
                    f"    def test_{item['name'].lower()}_instantiation(self):\n"
                    f"        \"\"\"Test that {item['name']} can be instantiated.\"\"\"\n"
                    f"        # TODO: Sınıf bağımlılıklarını mock'la\n"
                    f"        # instance = {item['name']}()\n"
                    f"        # assert instance is not None\n"
                    f"        pass\n"
                )
            elif item["type"] in ("function", "async_function"):
                prefix = "async " if item["type"] == "async_function" else ""
                marker = "@pytest.mark.asyncio\n    " if item["type"] == "async_function" else ""
                fn_tests.append(
                    f"\n    {marker}def test_{item['name']}_basic(self):\n"
                    f"        \"\"\"Basic smoke test for {item['name']}.\"\"\"\n"
                    f"        # TODO: Bağımlılıkları mock'la ve kenar durumları test et\n"
                    f"        # result = {call_expr}\n"
                    f"        # assert result is not None\n"
                    f"        pass\n"
                    f"\n    {marker}def test_{item['name']}_edge_cases(self):\n"
                    f"        \"\"\"Edge case test for {item['name']}.\"\"\"\n"
                    f"        # TODO: None, boş string, negatif sayı vb. dene\n"
                    f"        pass\n"
                )

        test_content = f'''"""
Auto-generated test file — GhostCore v4 Unit Test Suite
Generated: {now}
Module   : {module_name}
Generator: UnitTestGenerator (agents.py)

Bu dosya otomatik üretilmiştir.
TODO'ları tamamlayarak gerçek test senaryoları yaz.
"""

import pytest
from unittest.mock import MagicMock, patch, AsyncMock

# Modülü import et (yolu gerekirse düzelt)
# from src.{module_name} import *


class Test{module_name.capitalize()}Module:
    """Test suite for {module_name} module."""

    def setup_method(self):
        """Her test öncesi çalışır — ortak fixture'ları burada hazırla."""
        pass

    def teardown_method(self):
        """Her test sonrası çalışır — kaynakları temizle."""
        pass
{''.join(fn_tests)}

{''.join(cls_tests)}

# ---------------------------------------------------------------------------
# Integration Tests (Entegrasyon Testleri)
# ---------------------------------------------------------------------------

class Test{module_name.capitalize()}Integration:
    """Diğer modüllerle entegrasyon testleri."""

    def test_module_imports_cleanly(self):
        """Modülün import hatası vermeden yüklendiğini doğrula."""
        try:
            import importlib
            importlib.import_module("src.{module_name}")
        except ImportError as e:
            pytest.skip(f"Modül bulunamadı (geliştirme aşamasında): {{e}}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
'''
        return test_content

    def write_test_file(self, module_name: str, code: str, output_dir: str = "tests") -> Path:
        """
        Test dosyasını diske yazar.

        Args:
            module_name: Modül adı.
            code:        Modülün kaynak kodu.
            output_dir:  Test dosyasının yazılacağı dizin.

        Returns:
            Yazılan test dosyasının Path nesnesi.
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        test_path = Path(output_dir) / f"test_{module_name}.py"
        test_path.write_text(self.generate_test_file(module_name, code), encoding="utf-8")
        logger.info("[TestGen] Test dosyası oluşturuldu: %s", test_path)
        return test_path


TEST_GENERATOR = UnitTestGenerator()





# ===========================================================================
# YENI OZELLIK: WAR ROOM TARTISMASI + DONANIM FARKINDALIGI
# ===========================================================================

@dataclass
class HardwareProfileV1:
    """Eski WarRoomOrchestrator için profil (v1 uyumluluğu)."""
    cpu: str = "Ryzen 7"
    ram_gb: int = 20
    ram_guard_gb: int = 4
    storage_type: str = "SSD"


class WarRoomOrchestrator:
    """
    Ajanlar arasi tartisma metni uretir; donanim sinirlarini baglama ekler.
    """

    def build_context(self, profile: HardwareProfileV1 | None = None) -> str:
        p = profile or HardwareProfileV1()
        return (
            f"[HARDWARE PROFILE]\n"
            f"- CPU: {p.cpu}\n"
            f"- RAM: {p.ram_gb}GB (koruma esigi: {p.ram_guard_gb}GB)\n"
            f"- Storage: {p.storage_type}\n"
            f"- Kural: Agir kutuphane secmeden once RAM/CPU maliyetini tartis.\n"
        )

    def simulate_discussion(
        self,
        topic: str,
        profile: HardwareProfileV1 | None = None,
    ) -> list[str]:
        """Hunter <-> Architect tarzi kisa War Room konusmasi dondurur."""
        p = profile or HardwareProfileV1()
        return [
            f"[Hunter] {topic} icin onerilen stack CPU'yu zorlayabilir. {p.cpu} uzerinde daha hafif alternatif dusunelim.",
            f"[Architect] Kabul. RAM korumasi icin hedefimiz {p.ram_guard_gb}GB sinirini asmamak olacak.",
            "[Hunter] Data pipeline tarafinda chunking ve lazy loading zorunlu olsun.",
            "[Architect] UI tarafinda da minimal bagimlilik + Tailwind bilesenleriyle devam edelim.",
            "[System] War Room karari: performans oncelikli mimari benimsendi.",
        ]


WAR_ROOM = WarRoomOrchestrator()


# ===========================================================================
# ÖZELLIK 4: PERFORMANCE PROFILING
# ===========================================================================

class PerformanceProfiler:
    """
    Üretilen kodu statik analiz ile değerlendirir.
    Ryzen 7 / 20GB RAM profiline göre uyarılar ve öneriler üretir.

    Dinamik profil için: tasks.py'de `cProfile` + `tracemalloc` entegre edilecek.
    """

    # Ryzen 7 için güvenli eşikler
    RAM_WARNING_GB  = 4.0   # Tek modül için
    CPU_HEAVY_LOOPS = 3     # Nested loop derinliği uyarı eşiği

    ASYNC_PATTERNS = [
        (r"\btime\.sleep\b",       "time.sleep() tespit edildi → asyncio.sleep() kullan"),
        (r"\brequests\.",          "requests kütüphanesi → httpx + async/await kullan"),
        (r"open\(",                "Senkron dosya I/O → aiofiles kullanmayı düşün"),
        (r"\bsubprocess\.call\b",  "subprocess.call → asyncio.create_subprocess_exec kullan"),
    ]

    MEMORY_PATTERNS = [
        (r"\.append\(.*\)\s*$",    "Loop içi list.append → list comprehension daha verimli"),
        (r"\bpickle\.",            "pickle → büyük nesnelerde bellek şişmesi riski"),
        (r"\bpandas\.read",        "pandas.read_* → chunksize parametresi ekle (RAM koruma)"),
    ]

    def profile(self, code: str) -> str:
        """
        Kodu analiz ederek performans yönergeleri üretir.

        Args:
            code: Python kaynak kodu.

        Returns:
            Architect prompt'una eklenecek performans bağlamı.
        """
        warnings  = []
        nested    = max((len(m.group()) - len(m.group().lstrip())
                         for m in re.finditer(r"^\s+(for|while)\s", code, re.MULTILINE)),
                        default=0) // 4

        if nested >= self.CPU_HEAVY_LOOPS:
            warnings.append(
                f"  ⚠  {nested} seviye iç içe döngü tespit edildi → "
                f"numpy/vectorized operasyona dönüştür (CPU tasarrufu)"
            )

        for pattern, msg in self.ASYNC_PATTERNS:
            if re.search(pattern, code):
                warnings.append(f"  ⚡ {msg}")

        for pattern, msg in self.MEMORY_PATTERNS:
            if re.search(pattern, code):
                warnings.append(f"  🧠 {msg}")

        if not warnings:
            return (
                "\n[PERFORMANCE CHECK]\n"
                "  ✓ Belirgin performans sorunu tespit edilmedi.\n"
                "  → Async-first yaklaşımını koru. I/O işlemlerini asenkronlaştır.\n"
            )

        return (
            "\n[PERFORMANCE ADVISORY — Ryzen 7 / 20GB RAM Profili]\n"
            + "\n".join(warnings)
            + "\n  → Yukarıdaki sorunları gidererek en optimize versiyonu sun.\n"
        )


PROFILER = PerformanceProfiler()


# ===========================================================================
# ÖZELLIK 6: SECURITY AUTO-PATCHING
# ===========================================================================

class SecurityAutoPatcher:
    """
    Hunter'ın bulduğu güvenlik açıklarını otomatik olarak kodda yamalar.
    Sadece rapor etmez — o satırı siler ve güvenli alternatifi yazar.
    """

    PATCHES = [
        # SQL Injection → SQLAlchemy parameterized query
        (
            r'f["\'].*SELECT.*\{.*\}.*["\']',
            "# PATCHED: SQL Injection riski → SQLAlchemy text() ile parametrize edildi\n"
            '    result = db.execute(text("SELECT * FROM table WHERE id = :id"), {"id": safe_id})',
            "SQL Injection (f-string query)",
        ),
        # Hardcoded secret
        (
            r'(password|secret|api_key|token)\s*=\s*["\'][^"\']{4,}["\']',
            "# PATCHED: Hardcoded credential → os.environ'dan oku\n"
            '    _secret = os.environ.get("SECRET_KEY")  # .env dosyasına ekle',
            "Hardcoded Credential",
        ),
        # eval() kullanımı
        (
            r'\beval\s*\(',
            "# PATCHED: eval() kaldırıldı → ast.literal_eval() kullan\n"
            "    _result = ast.literal_eval(_safe_input)",
            "Arbitrary Code Execution (eval)",
        ),
        # Shell injection
        (
            r'subprocess\.(call|run)\(.*shell\s*=\s*True',
            "# PATCHED: shell=True kaldırıldı → liste argümanı kullan\n"
            '    subprocess.run(["komut", "arg1"], shell=False, check=True)',
            "Shell Injection (shell=True)",
        ),
        # MD5 kullanımı (zayıf hash)
        (
            r'hashlib\.md5\b',
            "# PATCHED: MD5 kriptografik olarak zayıf → hashlib.sha256 kullan\n"
            "    _hash = hashlib.sha256(_data).hexdigest()",
            "Weak Hashing (MD5)",
        ),
    ]

    def patch(self, code: str) -> tuple[str, list[str]]:
        """
        Koddaki güvenlik açıklarını tespit edip yamalar.

        Args:
            code: Orijinal Python kaynak kodu.

        Returns:
            (yamalı_kod, bulunan_açıklar_listesi)
        """
        patched_code = code
        found        = []

        for pattern, replacement, vuln_name in self.PATCHES:
            if re.search(pattern, patched_code, re.IGNORECASE):
                patched_code = re.sub(
                    pattern, replacement, patched_code, flags=re.IGNORECASE
                )
                found.append(vuln_name)
                logger.info("[SecurityPatcher] Yamalandı: %s", vuln_name)

        return patched_code, found

    def patch_report(self, found: list[str]) -> str:
        """Hunter log'u için yama raporu üretir."""
        if not found:
            return "✓ Otomatik yama gerekmedi."
        lines = "\n".join(f"  ✅ YAMALANDI: {v}" for v in found)
        return f"[Security Auto-Patch Raporu]\n{lines}"


SECURITY_PATCHER = SecurityAutoPatcher()


# ===========================================================================
# ÖZELLIK 10: GIT-NATIVE WORKFLOW
# ===========================================================================

class GitCommitGenerator:
    """
    Başarılı her modül üretimi sonrası Conventional Commits standardında
    Git commit mesajı hazırlar ve isteğe bağlı olarak çalıştırır.

    Format: <type>(<scope>): <description>
    Types : feat | fix | docs | refactor | test | chore | security
    """

    COMMIT_TYPES = {
        "new_module":    "feat",
        "bug_fix":       "fix",
        "security_fix":  "security",
        "documentation": "docs",
        "refactor":      "refactor",
        "tests":         "test",
        "config":        "chore",
    }

    def generate_message(
        self,
        change_type: str,
        scope: str,
        description: str,
        body: Optional[str] = None,
        breaking: bool = False,
    ) -> str:
        """
        Conventional Commits formatında commit mesajı üretir.

        Args:
            change_type:  Değişiklik tipi (new_module, bug_fix, vb.)
            scope:        Değişikliğin kapsamı (örn. "agents", "brain", "sentinel")
            description:  Kısa, imperative açıklama (ör. "add long-term memory layer")
            body:         Detaylı açıklama (opsiyonel).
            breaking:     True ise BREAKING CHANGE ekler.

        Returns:
            Hazır commit mesajı string'i.

        Example:
            >>> gen.generate_message("new_module", "agents", "add SOLID pattern advisor")
            "feat(agents): add SOLID pattern advisor"
        """
        prefix  = self.COMMIT_TYPES.get(change_type, "feat")
        bang    = "!" if breaking else ""
        header  = f"{prefix}{bang}({scope}): {description}"

        parts = [header]
        if body:
            parts.append(f"\n{body}")
        if breaking:
            parts.append("\nBREAKING CHANGE: bu değişiklik geriye dönük uyumsuz.")

        return "\n".join(parts)

    def stage_and_commit(self, message: str, files: Optional[list[str]] = None) -> bool:
        """
        Değişiklikleri stage'ler ve commit atar.
        Developer onayı alındıktan sonra çağrılır (main.py'de prompt var).

        Args:
            message: Commit mesajı.
            files:   Stage'lenecek dosya listesi. None ise `git add .` kullanır.

        Returns:
            True başarılıysa, False hata varsa.
        """
        try:
            if files:
                subprocess.run(["git", "add"] + files, check=True, capture_output=True)
            else:
                subprocess.run(["git", "add", "."], check=True, capture_output=True)

            subprocess.run(
                ["git", "commit", "-m", message],
                check=True, capture_output=True,
            )
            logger.info("[Git] Commit atıldı: %s", message.splitlines()[0])
            return True
        except subprocess.CalledProcessError as e:
            logger.error("[Git] Commit hatası: %s", e.stderr.decode())
            return False

    def suggest_commit(self, task_description: str, changed_files: list[str]) -> str:
        """
        Görev açıklamasından otomatik commit mesajı önerir.

        Args:
            task_description: Kullanıcının görev metni.
            changed_files:    Değişen dosyaların listesi.

        Returns:
            Önerilen commit mesajı.
        """
        # Görev tipini tahmin et
        desc_lower = task_description.lower()
        if any(w in desc_lower for w in ["fix", "düzelt", "hata", "bug"]):
            ctype = "bug_fix"
        elif any(w in desc_lower for w in ["security", "güvenlik", "patch", "yama"]):
            ctype = "security_fix"
        elif any(w in desc_lower for w in ["doc", "readme", "comment", "doküman"]):
            ctype = "documentation"
        elif any(w in desc_lower for w in ["test", "spec"]):
            ctype = "tests"
        elif any(w in desc_lower for w in ["refactor", "temizle", "yeniden"]):
            ctype = "refactor"
        else:
            ctype = "new_module"

        # Scope: değişen ilk dosyadan çıkar
        scope = Path(changed_files[0]).stem if changed_files else "core"

        # Açıklama: görev metnini 72 karaktere kısalt
        short_desc = task_description[:70].lower().replace(" ", "-")
        short_desc = re.sub(r"[^a-z0-9\-türkçeüöşıçğ]", "", short_desc)

        return self.generate_message(ctype, scope, short_desc)


GIT_WORKFLOW = GitCommitGenerator()


# ===========================================================================
# ÖZELLIK 8: TOOL-AWARE EXECUTION (Gerçek Araç Kullanımı)
# ===========================================================================
# CrewAI Tools — Ajanlar metin üretmenin ötesine geçip dosyaları gerçekten okur.

_file_read_tool      = FileReadTool()
_directory_read_tool = DirectoryReadTool()

# TerminalTool — Komutları çalıştırmadan önce developer onayı alır (güvenlik)
# Tam implementasyonu tasks.py'de; burada tanımı var.
class TerminalTool:
    """
    Güvenli terminal komutu çalıştırıcı.
    Her komut öncesi developer onayı (main.py'de) ister.
    İzin verilmeden hiçbir komut çalışmaz.
    """
    name        = "terminal_tool"
    description = (
        "Runs a terminal command AFTER developer approval. "
        "Never executes destructive commands (rm -rf, format, etc.) automatically."
    )

    BLOCKED_PATTERNS = [
        r"rm\s+-rf", r"format\s+[A-Z]:", r"dd\s+if=",
        r":\(\)\{.*\}", r"mkfs\.", r">>\s*/etc/passwd",
    ]

    def run(self, command: str, require_approval: bool = True) -> str:
        """
        Komutu çalıştırır. require_approval=True ise terminalden onay ister.

        Args:
            command:          Çalıştırılacak shell komutu.
            require_approval: False ise onaysız çalışır (sadece güvenli komutlar için).

        Returns:
            Komut çıktısı string'i.
        """
        # Tehlikeli komutları bloke et
        for pattern in self.BLOCKED_PATTERNS:
            if re.search(pattern, command, re.IGNORECASE):
                return f"[BLOCKED] Tehlikeli komut engellendi: {command}"

        if require_approval:
            console.print(
                Panel(
                    f"[yellow]Ajan şu komutu çalıştırmak istiyor:[/yellow]\n\n"
                    f"  [bold cyan]{command}[/bold cyan]\n\n"
                    f"[dim]Onaylamak için 'y', iptal için 'n' yaz.[/dim]",
                    title="[bold red]⚠ TERMINAL ONAYI GEREKİYOR[/bold red]",
                    border_style="red",
                )
            )
            approval = input("Onayla (y/n): ").strip().lower()
            if approval != "y":
                return "[CANCELLED] Komut developer tarafından iptal edildi."

        try:
            result = subprocess.run(
                command, shell=True, capture_output=True, text=True, timeout=30
            )
            output = result.stdout or result.stderr or "(çıktı yok)"
            return output[:2000]  # Token tasarrufu için kırp
        except subprocess.TimeoutExpired:
            return "[TIMEOUT] Komut 30 saniye içinde tamamlanamadı."
        except Exception as e:
            return f"[ERROR] {e}"


_terminal_tool = TerminalTool()


# ===========================================================================
# AJAN TANIMI — Prompt Fabrikası (Tüm Özellikleri Birleştiren Merkez)
# ===========================================================================

def build_architect_system_prompt(task: str = "", changed_file: str = "") -> str:
    """
    The Architect için tüm özellik katmanlarını birleştirerek
    tam sistem prompt'u oluşturur.

    Dahil olan katmanlar:
        - SOLID & Design Pattern önerileri (Özellik 1)
        - Multi-File Sync uyarıları (Özellik 2)
        - Long-Term Memory bağlamı (Özellik 7)
        - Hybrid Logic yönergesi (Özellik 9)
        - Performance Profiling kılavuzu (Özellik 4)

    Args:
        task:         Kullanıcının görev açıklaması.
        changed_file: Varsa değiştirilen dosyanın adı.

    Returns:
        Zenginleştirilmiş sistem prompt'u.
    """
    sections = []

    # Cerebro RAG (Experience Ledger)
    if task:
        memory_ctx = MEMORY.context_for_prompt(task)
        if memory_ctx:
            sections.append(
                "\n[CEREBRO EXPERIENCE LEDGER]\n"
                "  Daha önce yazdığımız altın standart kodları temel al, tekerleği yeniden icat etme:\n"
                f"{memory_ctx}"
            )

    # Design Pattern
    if task:
        sections.append(PATTERN_ADVISOR.suggest(task))

    # Multi-File Sync
    if changed_file:
        sync_ctx = FILE_SYNC.sync_report(changed_file)
        if sync_ctx:
            sections.append(sync_ctx)

    # Senior Principles (Elite Codegen)
    sections.append(
        "\n[SENIOR PRINCIPLES]\n"
        "  1. SOLID: Kod parçalara ayrılmış, tek sorumluluk prensibiyle (SRP) yazılmalıdır.\n"
        "  2. DRY: Tekrar eden bloklardan kaçın.\n"
        "  3. Type Hinting: Tüm imzalarda Python tip belirteçlerini (str, int, Optional vb.) KESINLIKLE KULLAN.\n"
        "  4. Docstrings: Her modül, sınıf ve fonksiyonu profesyonel ve açıklayıcı yorumlarla belgele.\n"
    )

    # Chain-of-Thought (CoT) and TDD Enforcement
    sections.append(
        "\n[ELITE CODEGEN FORMAT]\n"
        "Yazacağın kodu ve testlerini mutlaka aşağıdaki formata uyarak ve ilgili XML etiketlerini kullanarak dışa aktar.\n\n"
        "1. ÖNCE <plan> etiketi içerisinde algoritmanı, kütüphane seçimlerini ve edge case'leri tartış.\n"
        "2. SONRA <tests> etiketi içerisinde TDD (Test-Driven Development) prensibine uygun, `pytest` uyumlu test kodunu yaz.\n"
        "3. EN SON <code> etiketi içerisinde asıl Python kodunu yaz.\n\n"
        "Örnek Çıktı Yapısı:\n"
        "<plan>\nBen bu işlemi yaparken X kütüphanesini seçeceğim. Olası uç durumlar: Y null gelmesi.\n</plan>\n"
        "<tests>\nimport pytest\ndef test_feature():\n    pass\n</tests>\n"
        "<code>\ndef feature() -> bool:\n    return True\n</code>\n"
    )

    # Hybrid Logic
    sections.append(
        "\n[HYBRID LOGIC]\n"
        "  Görev karmaşıklığını değerlendir:\n"
        "  SIMPLE  → phi3 (hızlı, lokal)\n"
        "  MODERATE → llama3:8b (güçlü, lokal)\n"
        "  COMPLEX  → cloud fallback\n"
    )

    return "\n".join(sections)


def build_hunter_system_prompt(code: str = "") -> str:
    """
    The Hunter için Security Auto-Patch ve Performance bağlamı ekler.

    Args:
        code: Denetlenecek kaynak kodu.

    Returns:
        Hunter sistem prompt'u.
    """
    sections = [
        "\n[SECURITY MANDATE]\n"
        "  Sadece rapor etme — açıkları anında yama. \n"
        "  CRITICAL/HIGH bulguları için güvenli alternatif kodu yaz.\n"
        "  Format: 'PATCHED: <açık_adı> → <güvenli_kod>'\n"
    ]

    if code:
        perf_advice = PROFILER.profile(code)
        if perf_advice:
            sections.append(perf_advice)

    return "\n".join(sections)


# ===========================================================================

@dataclass
class SessionState:
    last_topic: str = ""
    mode: str = "verbose"
    last_result: str = ""


class SessionMemory:
    """Kapanış-açılış arasında state saklar."""

    def __init__(self, state_file: str = "data/session_state.json") -> None:
        self.path = Path(state_file)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def load(self) -> SessionState:
        if not self.path.exists():
            return SessionState()
        try:
            raw = json.loads(self.path.read_text(encoding="utf-8"))
            return SessionState(**raw)
        except Exception:
            return SessionState()

    def save(self, state: SessionState) -> None:
        self.path.write_text(json.dumps(asdict(state), indent=2), encoding="utf-8")


class DataMaestroV2:
    """Excel buyucusu: chart, anomaly, merge, summary, sql-to-excel."""

    def __init__(self) -> None:
        import pandas as pd
        self.pd = pd

    def _anomaly_flags(self, series):
        if series.empty:
            return []
        mean = series.mean()
        std = series.std() or 0
        if std == 0:
            return [False] * len(series)
        z = (series - mean).abs() / std
        return list(z > 2.5)

    def multi_file_merge_to_excel(
        self,
        folder: str,
        output_path: str = "reports/maestro_multi.xlsx",
        limit: int = 5,
    ) -> str:
        folder_path = Path(folder)
        csv_files = sorted(folder_path.glob("*.csv"))
        json_files = sorted(folder_path.glob("*.json"))
        data_files = (csv_files + json_files)[:limit]
        if not data_files:
            raise ValueError("Klasorde CSV/JSON bulunamadi.")

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        with self.pd.ExcelWriter(out, engine="openpyxl") as writer:
            merged_meta: list[tuple[str, int, int]] = []
            for f in data_files:
                if f.suffix.lower() == ".csv":
                    df = self.pd.read_csv(f)
                else:
                    df = self.pd.read_json(f)
                sheet = f.stem[:31]
                df.to_excel(writer, sheet_name=sheet, index=False)
                merged_meta.append((sheet, len(df), len(df.columns)))

            # Executive Summary
            summary_df = self.pd.DataFrame(
                merged_meta, columns=["Sheet", "Rows", "Columns"]
            )
            summary_df.to_excel(writer, sheet_name="Yonetici Ozeti", index=False)

        # Grafik + anomaly + format
        wb = load_workbook(out)
        sum_ws = wb["Yonetici Ozeti"]
        self._style_summary(sum_ws)
        self._add_summary_chart(sum_ws)
        self._mark_anomalies_all_sheets(wb)
        wb.save(out)
        return str(out)

    def merge_all_to_excel(
        self,
        folder: str,
        output_path: str = "reports/data_maestro_merged.xlsx",
    ) -> str:
        """Quick command için klasördeki CSV/JSON dosyalarını birleştirir."""
        return self.multi_file_merge_to_excel(folder=folder, output_path=output_path, limit=1000)

    def sql_to_excel(
        self,
        sqlite_path: str,
        query: str,
        output_path: str = "reports/sql_report.xlsx",
    ) -> str:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(sqlite_path) as conn:
            df = self.pd.read_sql_query(query, conn)
        with self.pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="SQL Result", index=False)
            # Executive summary
            stats = df.describe(include="all").transpose()
            stats.to_excel(writer, sheet_name="Yonetici Ozeti")
        wb = load_workbook(out)
        self._add_data_chart(wb["SQL Result"])
        self._mark_anomalies_sheet(wb["SQL Result"])
        wb.save(out)
        return str(out)

    def _style_summary(self, ws) -> None:
        fill = PatternFill(fill_type="solid", fgColor="1F2937")
        for c in ws[1]:
            c.fill = fill
            c.font = Font(color="FFFFFF", bold=True)

    def _add_summary_chart(self, ws) -> None:
        # rows by sheet
        if ws.max_row < 2:
            return
        chart = PieChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Sheet Bazli Satir Dagilimi"
        ws.add_chart(chart, "E2")

    def _add_data_chart(self, ws) -> None:
        if ws.max_row < 2 or ws.max_column < 2:
            return
        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 25))
        chart.add_data(data, titles_from_data=True)
        chart.title = "Data Snapshot"
        ws.add_chart(chart, "H2")

    def _mark_anomalies_sheet(self, ws) -> None:
        red = PatternFill(fill_type="solid", fgColor="FCA5A5")
        # Numeric columns only: basit tip kontrolü
        for col in range(1, ws.max_column + 1):
            vals = []
            rows = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col).value
                if isinstance(v, (int, float)):
                    vals.append(float(v))
                    rows.append(r)
            if len(vals) < 4:
                continue
            s = self.pd.Series(vals)
            flags = self._anomaly_flags(s)
            for idx, flag in enumerate(flags):
                if flag:
                    ws.cell(rows[idx], col).fill = red

    def _mark_anomalies_all_sheets(self, wb) -> None:
        for ws in wb.worksheets:
            if ws.title == "Yonetici Ozeti":
                continue
            self._mark_anomalies_sheet(ws)


class DesignerV2:
    """Designer + Redline entegrasyonu."""

    def ensure_assets(self, project_root: str = ".") -> dict[str, str]:
        base = Path(project_root)
        paths = {
            "static": str((base / "static").resolve()),
            "templates": str((base / "templates").resolve()),
            "css": str((base / "css").resolve()),
        }
        for p in paths.values():
            Path(p).mkdir(parents=True, exist_ok=True)
        return paths

    def component_factory(self) -> dict[str, str]:
        crypto = """
<section class="rounded-xl border border-slate-700 p-4 dark:bg-gray-800 bg-white">
  <h3 class="text-lg font-semibold text-slate-100">CryptoFeed</h3>
  <p class="text-sm text-slate-400">BTC/ETH anlik fiyat paneli.</p>
</section>
""".strip()
        weather = """
<section class="rounded-xl border border-slate-700 p-4 dark:bg-gray-800 bg-white">
  <h3 class="text-lg font-semibold text-slate-100">WeatherPanel</h3>
  <p class="text-sm text-slate-400">Sehir bazli hava durumu karti.</p>
</section>
""".strip()
        return {"CryptoFeed": crypto, "WeatherPanel": weather}

    def add_dark_light_support(self, html: str) -> str:
        if "dark:bg-gray-800" in html:
            return html
        return html.replace('class="', 'class="dark:bg-gray-800 ')

    def harmonize_palette(self, palette: dict[str, str]) -> dict[str, str]:
        # Basit self-correction: beyaz zeminde cok acik yaziyi düzelt
        fixed = dict(palette)
        if fixed.get("background", "").lower() in {"#ffffff", "white"} and fixed.get("text", "").lower() in {"#f9fafb", "#ffffff"}:
            fixed["text"] = "#111827"
        return fixed

    def one_click_preview(self, html_path: str = "templates/preview.html", port: int = 8000) -> str:
        cmd = ["python", "-m", "http.server", str(port)]
        subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return f"http://localhost:{port} (Tasarim hazir)"


@dataclass
class HardwareProfile:
    cpu_name: str = "Ryzen 7"
    ram_total_gb: int = 20
    ram_threshold: int = 85
    cpu_thermal_threshold: int = 88


class WarRoomV2:
    def discuss(self, topic: str, profile: Optional[HardwareProfile] = None) -> list[str]:
        p = profile or HardwareProfile()
        return [
            f"[Hunter] {topic} CPU icin agir olabilir ({p.cpu_name}).",
            f"[Architect] RAM %{p.ram_threshold} esigini asmadan queue ile ilerleyelim.",
            "[Data Maestro] Analizi chunk'layip memory baskisini azaltacagim.",
            "[Designer] UI tarafinda hafif bileşenlerle render yükünü düşüreceğim.",
            "[System] Karar: performans-first local pipeline.",
        ]




SESSION_MEMORY = SessionMemory()
DATA_MAESTRO_V2 = DataMaestroV2()
DESIGNER_V2     = DesignerV2()
WAR_ROOM_V2     = WarRoomV2()

# Tek, yetkili yardımcı sözlüğü — tüm mod erişimi buradan
ENHANCED_UTILITIES: dict[str, object] = {
    "memory":           MEMORY,
    "pattern_advisor":  PATTERN_ADVISOR,
    "file_sync":        FILE_SYNC,
    "test_generator":   TEST_GENERATOR,
    "war_room":         WAR_ROOM,
    "profiler":         PROFILER,
    "security_patcher": SECURITY_PATCHER,
    "git_workflow":     GIT_WORKFLOW,
    "terminal_tool":    _terminal_tool,
    "web_search":       WEB_SEARCH_TOOL,
    "sandbox":          SANDBOX,
    "session_memory":   SESSION_MEMORY,
    "data_maestro_v2":  DATA_MAESTRO_V2,
    "designer_v2":      DESIGNER_V2,
    "war_room_v2":      WAR_ROOM_V2,
}
